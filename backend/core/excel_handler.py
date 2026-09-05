"""
Módulo para criação e manipulação de planilhas Excel no projeto MARIA.
Usa pandas + openpyxl para geração e manipulação de arquivos .xlsx.

v4.3.x: retornos estruturados preparados para visualização em tempo real no frontend.
"""

import os
import logging
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from backend.core.file_utils import garantir_pasta_arquivos, gerar_nome_unico, sanitizar_nome_arquivo
from backend.core.config import PASTA_ARQUIVOS_GERADOS, get_max_linhas_por_chamada

logger = logging.getLogger(__name__)


def _aplicar_estilo_cabecalho(caminho: str, linha_cabecalho: int, num_colunas: int) -> None:
    """Aplica negrito nos cabeçalhos e ajusta largura das colunas via openpyxl."""
    wb = load_workbook(caminho)
    ws = wb.active
    font_negrito = Font(bold=True)
    for col_idx in range(1, num_colunas + 1):
        cell = ws.cell(row=linha_cabecalho, column=col_idx)
        cell.font = font_negrito
        col_letter = get_column_letter(col_idx)
        valor = str(cell.value or "")
        ws.column_dimensions[col_letter].width = max(len(valor) + 2, 10)
    wb.save(caminho)
    wb.close()


def criar_planilha_real(
    nome_arquivo: str,
    colunas: list[str],
    descricao: str = "",
    linhas: list[dict] | None = None,
) -> str:
    """
    Cria uma planilha Excel com as colunas especificadas e, opcionalmente, linhas de dados.

    Args:
        nome_arquivo: Nome do arquivo (com ou sem .xlsx — normalizado internamente)
        colunas: Lista de nomes das colunas
        descricao: Descrição opcional exibida como título na primeira linha
        linhas: Lista de dicts com dados. Chaves devem corresponder a `colunas`.
                Colunas ausentes ficam vazias; chaves extras são ignoradas.
                Parâmetro opcional — None mantém comportamento anterior (só cabeçalho).
                O número de linhas é limitado automaticamente por get_max_linhas_por_chamada().

    Returns:
        Caminho absoluto do arquivo criado

    # v4.3.x: retornar dict {"caminho": str, "linhas_escritas": int, "colunas": list}
    #         para atualização em tempo real no frontend TableView.

    Raises:
        PermissionError: Sem permissão de escrita
        OSError: Erro de disco
    """
    try:
        if nome_arquivo.endswith(".xlsx"):
            nome_arquivo = nome_arquivo[:-5]

        nome_final = gerar_nome_unico(nome_arquivo, ".xlsx")
        pasta_absoluta = garantir_pasta_arquivos()
        caminho_completo = os.path.join(pasta_absoluta, nome_final)

        # Aplicar limite de linhas por modelo (transparente ao usuário)
        linhas_dados = linhas or []
        limite = get_max_linhas_por_chamada()
        if len(linhas_dados) > limite:
            logger.warning(
                "criar_planilha: %d linhas recebidas, limite do modelo é %d. Truncando.",
                len(linhas_dados), limite,
            )
            linhas_dados = linhas_dados[:limite]

        # Montar DataFrame com as colunas definidas
        # Colunas ausentes nas linhas ficam NaN → substituídas por string vazia
        df = pd.DataFrame(linhas_dados, columns=colunas)
        df = df.reindex(columns=colunas)
        df = df.fillna("")

        # Escrever arquivo
        startrow = 0
        if descricao:
            # Linha 1: descrição (via openpyxl após escrita do pandas)
            startrow = 2  # pandas escreve a partir da linha 3 (0-indexed: 2)

        with pd.ExcelWriter(caminho_completo, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Dados", index=False, startrow=startrow)

        # Adicionar descrição e aplicar estilos via openpyxl
        if descricao:
            wb = load_workbook(caminho_completo)
            ws = wb.active
            ws.cell(row=1, column=1, value=descricao)
            if len(colunas) > 1:
                ws.merge_cells(f"A1:{get_column_letter(len(colunas))}1")
            wb.save(caminho_completo)
            wb.close()

        linha_cabecalho = 3 if descricao else 1
        _aplicar_estilo_cabecalho(caminho_completo, linha_cabecalho, len(colunas))

        linhas_escritas = len(df)
        logger.info("Planilha criada: %s (%d linhas)", caminho_completo, linhas_escritas)
        return caminho_completo

    except PermissionError as e:
        logger.error("Permissão negada ao criar planilha: %s", e)
        raise PermissionError(
            f"Não foi possível salvar o arquivo. Verifique as permissões da pasta '{PASTA_ARQUIVOS_GERADOS}'."
        ) from e
    except OSError as e:
        logger.error("Erro de disco ao criar planilha: %s", e)
        raise OSError(
            "Não foi possível salvar o arquivo. Verifique se há espaço em disco disponível."
        ) from e
    except Exception as e:
        logger.error("Erro inesperado ao criar planilha: %s", e)
        raise


def editar_planilha_real(
    nome_arquivo: str,
    colunas: list[str],
    linhas: list[dict] | None = None,
    descricao: str = "",
) -> str:
    """
    Sobrescreve uma planilha existente com novas colunas e, opcionalmente, linhas de dados.

    Args:
        nome_arquivo: Nome exato do arquivo sem extensão
        colunas: Nova lista de nomes de colunas
        linhas: Lista de dicts com dados. Mesmas regras de criar_planilha_real().
                Parâmetro opcional — None sobrescreve com estrutura vazia.
        descricao: Descrição opcional

    Returns:
        Caminho absoluto do arquivo sobrescrito

    # v4.3.x: retornar dict {"caminho": str, "linhas_escritas": int, "colunas": list}
    #         para atualização em tempo real no frontend TableView.

    Raises:
        ValueError: Arquivo não encontrado
        PermissionError: Sem permissão de escrita
        OSError: Erro de disco
    """
    try:
        if nome_arquivo.endswith(".xlsx"):
            nome_arquivo = nome_arquivo[:-5]

        nome_seguro = sanitizar_nome_arquivo(nome_arquivo)
        pasta_absoluta = garantir_pasta_arquivos()
        caminho_completo = os.path.join(pasta_absoluta, f"{nome_seguro}.xlsx")

        if not os.path.exists(caminho_completo):
            raise ValueError(
                f"Arquivo '{nome_seguro}.xlsx' não encontrado na pasta de arquivos gerados."
            )

        # Aplicar limite de linhas por modelo
        linhas_dados = linhas or []
        limite = get_max_linhas_por_chamada()
        if len(linhas_dados) > limite:
            logger.warning(
                "editar_planilha: %d linhas recebidas, limite do modelo é %d. Truncando.",
                len(linhas_dados), limite,
            )
            linhas_dados = linhas_dados[:limite]

        df = pd.DataFrame(linhas_dados, columns=colunas)
        df = df.reindex(columns=colunas)
        df = df.fillna("")

        startrow = 2 if descricao else 0

        with pd.ExcelWriter(caminho_completo, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Dados", index=False, startrow=startrow)

        if descricao:
            wb = load_workbook(caminho_completo)
            ws = wb.active
            ws.cell(row=1, column=1, value=descricao)
            if len(colunas) > 1:
                ws.merge_cells(f"A1:{get_column_letter(len(colunas))}1")
            wb.save(caminho_completo)
            wb.close()

        linha_cabecalho = 3 if descricao else 1
        _aplicar_estilo_cabecalho(caminho_completo, linha_cabecalho, len(colunas))

        linhas_escritas = len(df)
        logger.info("Planilha sobrescrita: %s (%d linhas)", caminho_completo, linhas_escritas)
        return caminho_completo

    except ValueError:
        raise
    except PermissionError as error:
        logger.error("Permissão negada ao editar planilha: %s", error)
        raise PermissionError(
            "Não foi possível salvar o arquivo. Verifique as permissões da pasta de arquivos gerados."
        ) from error
    except OSError as error:
        logger.error("Erro de disco ao editar planilha: %s", error)
        raise OSError(
            "Não foi possível salvar o arquivo. Verifique se há espaço em disco disponível."
        ) from error
    except Exception as error:
        logger.error("Erro inesperado ao editar planilha: %s", error)
        raise


def ler_planilha_resumo(caminho: str) -> str:
    """
    Lê uma planilha Excel e retorna um resumo textual básico.
    Mantém openpyxl (não usa pandas) — comportamento inalterado.

    Args:
        caminho: Caminho completo para o arquivo Excel

    Returns:
        String com resumo da planilha

    Raises:
        FileNotFoundError: Arquivo não encontrado
        ValueError: Arquivo inválido
    """
    try:
        if not os.path.exists(caminho):
            raise FileNotFoundError(f"Arquivo não encontrado: {caminho}")

        wb = load_workbook(caminho, read_only=True, data_only=True)
        ws = wb.active

        max_row = ws.max_row
        max_col = ws.max_column

        cabecalhos = []
        for col in range(1, min(max_col + 1, 10)):
            cell = ws.cell(row=1, column=col)
            cabecalhos.append(str(cell.value) if cell.value else f"Col{col}")

        amostra_dados = []
        for row in range(2, min(max_row + 1, 5)):
            linha_dados = []
            for col in range(1, min(max_col + 1, 5)):
                cell = ws.cell(row=row, column=col)
                linha_dados.append(str(cell.value) if cell.value else "")
            amostra_dados.append(" | ".join(linha_dados))

        wb.close()

        resumo = [
            f"Planilha: {os.path.basename(caminho)}",
            f"Linhas totais: {max_row - 1} (excluindo cabeçalho)",
            f"Colunas totais: {max_col}",
            f"Cabeçalhos: {', '.join(cabecalhos[:10])}",
        ]
        if amostra_dados:
            resumo.append("Amostra de dados (primeiras 3 linhas):")
            for i, linha in enumerate(amostra_dados, 1):
                resumo.append(f"  Linha {i}: {linha}")

        return "\n".join(resumo)

    except FileNotFoundError:
        raise
    except Exception as error:
        logger.error("Erro ao ler planilha: %s", error)
        raise ValueError(f"Não foi possível ler a planilha: {error}")
